import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload
from models.journal_model import Journal
from models.teacher_model import Teachers
from models.group_model import Groups
from models.subject_model import Subjects
from models.student_model import Students
from models.teacher_subject import TeacherSubject
from models.grade_model import Grade
from models.lesson_model import Lesson
from schemas.journals import JournalFullResponse, JournalListResponse, \
    LessonResponse, GroupShort, SubjectShort, TeacherShort, StudentShort
from schemas.grades import GradeResponse


class JournalService:

    @staticmethod
    async def create_journal(
            db: AsyncSession,
            group_id: int,
            subject_id: int,
            teacher_id: int,
            assistant_id: int | None = None,
    ) -> Journal:
        group = await db.get(Groups, group_id)
        if group is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Group not found")

        subject = await db.get(Subjects, subject_id)
        if subject is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Subject not found")

        teacher = await db.get(Teachers, teacher_id)
        if teacher is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Teacher not found")

        ts_res = await db.execute(
            select(TeacherSubject).where(
                TeacherSubject.teacher_id == teacher_id,
                TeacherSubject.subject_id == subject_id,
            )
        )
        if ts_res.scalar_one_or_none() is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Teacher is not assigned to this subject",
            )

        if assistant_id is not None:
            assistant = await db.get(Teachers, assistant_id)
            if assistant is None:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                    detail="Assistant not found")

        existing = await db.execute(
            select(Journal).where(
                Journal.group_id == group_id,
                Journal.subject_id == subject_id,
            )
        )
        if existing.scalar_one_or_none() is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Journal for this group and subject already exists",
            )

        journal = Journal(
            group_id=group_id,
            subject_id=subject_id,
            teacher_id=teacher_id,
            assistant_id=assistant_id,
        )
        db.add(journal)
        await db.commit()
        await db.refresh(journal)
        return journal

    @staticmethod
    async def get_journals(db: AsyncSession):
        result = await db.execute(
            select(Journal).options(
                joinedload(Journal.subject),
                joinedload(Journal.group),
                joinedload(Journal.teacher),
            )
        )
        journals = result.unique().scalars().all()

        # Compute last_updated per journal: max(updated_at) from lessons and grades
        lesson_max_res = await db.execute(
            select(Lesson.journal_id, func.max(Lesson.updated_at).label("max_updated"))
            .group_by(Lesson.journal_id)
        )
        lesson_max = {row.journal_id: row.max_updated for row in lesson_max_res}

        grade_max_res = await db.execute(
            select(Lesson.journal_id, func.max(Grade.updated_at).label("max_updated"))
            .join(Grade, Grade.lesson_id == Lesson.id)
            .group_by(Lesson.journal_id)
        )
        grade_max = {row.journal_id: row.max_updated for row in grade_max_res}

        def get_last_updated(journal_id):
            vals = [v for v in [lesson_max.get(journal_id), grade_max.get(journal_id)] if v is not None]
            return max(vals) if vals else None

        grouped = {}
        for journal in journals:
            subject_name = journal.subject.name
            if subject_name not in grouped:
                grouped[subject_name] = []
            group = journal.group
            grouped[subject_name].append({
                "journal_id": journal.id,
                "group_id": group.id,
                "name": group.name,
                "course_number": group.course_number,
                "teacher_name": journal.teacher.name if journal.teacher else None,
                "last_updated": get_last_updated(journal.id),
            })

        return [
            {"subject": subject, "groups": groups}
            for subject, groups in grouped.items()
        ]

    @staticmethod
    async def get_journal_by_id(db: AsyncSession,
                                journal_id: int) -> JournalFullResponse:
        journal = await db.get(Journal, journal_id)
        if journal is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Journal not found")

        students_res = await db.execute(
            select(Students)
            .where(Students.group_id == journal.group_id)
            .order_by(Students.name)
        )
        students = students_res.scalars().all()

        lesson_ids = [l.id for l in journal.lessons]
        grades = []
        if lesson_ids:
            grades_res = await db.execute(
                select(Grade).where(Grade.lesson_id.in_(lesson_ids))
            )
            grades = grades_res.scalars().all()

        return JournalFullResponse(
            id=journal.id,
            group=GroupShort(journal_id=journal.id, id=journal.group.id, name=journal.group.name,
                             course_number=journal.group.course_number),
            subject=SubjectShort(id=journal.subject.id,
                                 name=journal.subject.name),
            teacher=TeacherShort(id=journal.teacher.id,
                                 name=journal.teacher.name),
            assistant=TeacherShort(id=journal.assistant.id,
                                   name=journal.assistant.name) if journal.assistant else None,
            lessons=[
                LessonResponse(
                    id=l.id,
                    date=l.date,
                    lesson_type=l.lesson_type,
                    order_index=l.order_index,
                    title=l.title,
                    description=l.description,
                    classroom_id=l.classroom_id,
                )
                for l in sorted(journal.lessons, key=lambda x: x.order_index)
            ],
            students=[StudentShort(id=s.id, name=s.name) for s in students],
            grades=[
                GradeResponse(
                    id=g.id,
                    lesson_id=g.lesson_id,
                    student_id=g.student_id,
                    value=g.value,
                    remark=g.remark,
                )
                for g in grades
            ],
        )

    @staticmethod
    async def delete_journal(db: AsyncSession, journal_id: int) -> None:
        journal = await db.get(Journal, journal_id)
        if journal is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Journal not found")
        await db.delete(journal)
        await db.commit()


    @staticmethod
    async def export_journal_to_excel(db: AsyncSession, journal_id: int) -> bytes:
        journal = await db.get(Journal, journal_id)
        if journal is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                                detail="Journal not found")

        students_res = await db.execute(
            select(Students)
            .where(Students.group_id == journal.group_id)
            .order_by(Students.name)
        )
        students = students_res.scalars().all()

        lesson_ids = [l.id for l in journal.lessons]
        grades = []
        if lesson_ids:
            grades_res = await db.execute(
                select(Grade).where(Grade.lesson_id.in_(lesson_ids))
            )
            grades = grades_res.scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{journal.subject.name} - {journal.group.name}"

        ws.cell(row=1, column=1, value="Ім\'я студента")
        for idx, lesson in enumerate(sorted(journal.lessons, key=lambda x: x.order_index)):
            ws.cell(row=1, column=idx + 2, value=f"{lesson.date} ({lesson.lesson_type})")

        for row_idx, student in enumerate(students, start=2):
            ws.cell(row=row_idx, column=1, value=student.name)
            for col_idx, lesson in enumerate(sorted(journal.lessons, key=lambda x: x.order_index), start=2):
                grade = next((g for g in grades if g.lesson_id == lesson.id and g.student_id == student.id), None)
                if grade:
                    ws.cell(row=row_idx, column=col_idx, value=grade.value)

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    async def get_my_journals(db: AsyncSession, teacher_id: int):
        result = await db.execute(
            select(Journal).where(
                (Journal.teacher_id == teacher_id) | (Journal.assistant_id == teacher_id)
            ).options(
                joinedload(Journal.subject),
                joinedload(Journal.group),
                joinedload(Journal.teacher),
            )
        )
        journals = result.unique().scalars().all()

        return [
            JournalListResponse(
                id=journal.id,
                subject_name=journal.subject.name,
                group_name=journal.group.name,
                teacher_name=journal.teacher.name if journal.teacher else None,
            )
            for journal in journals
        ]


journal_service = JournalService()
